import openpyxl

# Write a function that reads in an excel file. It must read data from a cel, parse that data with a space as a parser and write it back to the cell and cells next to it
# The function must be able to read in a file and write to a file
# It must be abled to read from multiple sheets in the excel file
def read_excel_file(file_name, sheet_number, cell_location):
    # Read in the excel file
    wb = openpyxl.load_workbook(file_name)
    # Read in the sheet
    sheet = wb.worksheets[sheet_number]
    # Read in the cell
    data = sheet.cell(row=cell_location[0], column=cell_location[1]).value
    # Parse the cell
    if data is not None:
        print(data)
        parsed_data = data.split(" ")
        # Write the parsed data to the cell and cells next to it
        for i in range(len(parsed_data)):
            row = cell_location[0]
            column = cell_location[1] + i
            sheet.cell(row=row, column=column).value = parsed_data[i]

        wb.save(file_name)

        return parsed_data
    else:
        return None

for i in range(5, 46):
    for k in range(1, 8):
        j = 2
        print(read_excel_file("Resultaten.xlsx", k, (i, j)))