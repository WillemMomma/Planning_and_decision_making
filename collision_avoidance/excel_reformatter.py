import openpyxl


def reformat_excel_file(file_name, sheet_number, cell_location):
    # Read in the excel file
    wb = openpyxl.load_workbook(file_name)
    # Read in the sheet
    sheet = wb.worksheets[sheet_number]
    # Read in the cell
    data = sheet.cell(row=cell_location[0], column=cell_location[1]).value
    # Parse the cell
    if data is not None:
        print(data)
        parsed_data = data.split(" ")
        # Write the parsed data to the cell and cells next to it
        for i in range(len(parsed_data)):
            row = cell_location[0]
            column = cell_location[1] + i
            sheet.cell(row=row, column=column).value = parsed_data[i]

        wb.save(file_name)

        return parsed_data
    else:
        return None


def write_excel_file(file_name, sheet_number, cell_location, input):
    # Read in the excel file
    wb = openpyxl.load_workbook(file_name)
    # Read in the sheet
    sheet = wb.worksheets[sheet_number]

    # Set row and col variable
    col = cell_location[1]
    for row in range(cell_location[0], cell_location[0]+10):

        # Read in the cell
        data = sheet.cell(row=row, column=col).value
        if data is not None:
            continue
        else:
            for index, ding in enumerate(input):
                sheet.cell(row=row, column=col+index).value = ding
                wb.save(file_name)

            return [row, col]

    return [-1, -1], None

def turn_to_number(file_name, sheet_number, cell_location):
    # Read in the excel file
    wb = openpyxl.load_workbook(file_name)
    # Read in the sheet
    sheet = wb.worksheets[sheet_number]

    # Set row and col variable
    for row in range(cell_location[0], cell_location[0]+10):
        for col in range(cell_location[1], cell_location[1] + 10):

            # Read in the cell
            try:
                data = sheet.cell(row=row, column=col).value
                data = float(data)
                sheet.cell(row=row, column=col).value = data
                wb.save(file_name)
            except:
                continue
